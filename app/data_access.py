from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
import csv
import re
import unicodedata

import bcrypt
from openpyxl import load_workbook

from .config import ACCESS_FILE, CLASSIFICATION_FILE, CLIENTS_FILE, RESPONSES_FILE


DAY_NAMES = {
    "seg": 0,
    "segunda": 0,
    "segunda-feira": 0,
    "ter": 1,
    "terca": 1,
    "terça": 1,
    "terca-feira": 1,
    "terça-feira": 1,
    "qua": 2,
    "quarta": 2,
    "quarta-feira": 2,
    "qui": 3,
    "quinta": 3,
    "quinta-feira": 3,
    "sex": 4,
    "sexta": 4,
    "sexta-feira": 4,
    "sab": 5,
    "sábado": 5,
    "sabado": 5,
    "dom": 6,
    "domingo": 6,
}


@dataclass
class AccessDecision:
    allowed: bool
    reason: str = ""


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char)).lower()


def digits_only(value: object) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def slugify(value: object) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "arquivo"


def workbook_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    headers_row = next(iterator, ())
    headers = [normalize_text(header).replace(" ", "_") for header in headers_row]
    rows: list[dict[str, object]] = []
    for values in iterator:
        if not any(value not in (None, "") for value in values):
            continue
        row = {
            headers[index] if index < len(headers) else f"coluna_{index + 1}": values[index]
            for index in range(len(values))
            if index < len(headers) and headers[index]
        }
        rows.append(row)
    workbook.close()
    return rows


def load_access_users() -> list[dict[str, object]]:
    if not ACCESS_FILE.exists():
        return []
    return workbook_rows(ACCESS_FILE)


def find_user(login: str) -> dict[str, object] | None:
    normalized_login = normalize_text(login)
    for row in load_access_users():
        if normalize_text(row.get("login")) == normalized_login:
            return row
    return None


def verify_password(raw_password: str, stored_password: object) -> bool:
    stored = str(stored_password or "").strip()
    if not stored:
        return False
    if stored.startswith("$2a$") or stored.startswith("$2b$") or stored.startswith("$2y$"):
        return bcrypt.checkpw(raw_password.encode("utf-8"), stored.encode("utf-8"))
    return raw_password == stored


def parse_date_token(token: str) -> date | None:
    token = token.strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(token, fmt).date()
        except ValueError:
            continue
    return None


def parse_time_token(token: str) -> time | None:
    token = token.strip().replace("h", ":")
    for fmt in ("%H:%M", "%H:%M:%S", "%H"):
        try:
            return datetime.strptime(token, fmt).time()
        except ValueError:
            continue
    return None


def evaluate_date_rule(value: object, current_date: date) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    if "-" in text and text.count("-") >= 2:
        parsed = parse_date_token(text)
        return parsed == current_date if parsed else False
    separator = None
    for candidate in (" a ", " até ", " to "):
        if candidate in normalize_text(text):
            separator = candidate
            break
    if separator:
        normalized = normalize_text(text)
        left, right = normalized.split(separator, 1)
        start = parse_date_token(left)
        end = parse_date_token(right)
        if start and end:
            return start <= current_date <= end
        return False
    exact = parse_date_token(text)
    return exact == current_date if exact else False


def evaluate_day_rule(value: object, current_date: date) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    chunks = [chunk.strip() for chunk in re.split(r"[,;|]", text) if chunk.strip()]
    weekday = current_date.weekday()
    for chunk in chunks:
        normalized = normalize_text(chunk)
        if normalized in DAY_NAMES and DAY_NAMES[normalized] == weekday:
            return True
        parsed_date = parse_date_token(chunk)
        if parsed_date == current_date:
            return True
    return False


def evaluate_time_rule(value: object, current_time: time) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    ranges = [chunk.strip() for chunk in re.split(r"[,;|]", text) if chunk.strip()]
    for chunk in ranges:
        if "-" not in chunk:
            exact = parse_time_token(chunk)
            if exact and current_time.hour == exact.hour and current_time.minute == exact.minute:
                return True
            continue
        start_raw, end_raw = chunk.split("-", 1)
        start = parse_time_token(start_raw)
        end = parse_time_token(end_raw)
        if not start or not end:
            continue
        if start <= current_time <= end:
            return True
    return False


def evaluate_access_window(user_row: dict[str, object], now: datetime) -> AccessDecision:
    active_value = normalize_text(user_row.get("ativo"))
    if active_value in {"n", "nao", "não", "0", "inativo"}:
        return AccessDecision(False, "Usuario desativado.")

    if user_row.get("data_inicio"):
        start = parse_date_token(str(user_row.get("data_inicio")))
        if start and now.date() < start:
            return AccessDecision(False, "Acesso ainda nao liberado para este usuario.")

    if user_row.get("data_fim"):
        end = parse_date_token(str(user_row.get("data_fim")))
        if end and now.date() > end:
            return AccessDecision(False, "Periodo de acesso encerrado para este usuario.")

    if user_row.get("data") and not evaluate_date_rule(user_row.get("data"), now.date()):
        return AccessDecision(False, "Acesso bloqueado para a data atual.")

    if not evaluate_day_rule(user_row.get("dia"), now.date()):
        return AccessDecision(False, "Acesso indisponivel para o dia atual.")

    if user_row.get("hora_inicio") or user_row.get("hora_fim"):
        start = parse_time_token(str(user_row.get("hora_inicio") or ""))
        end = parse_time_token(str(user_row.get("hora_fim") or ""))
        if start and end and not (start <= now.time() <= end):
            return AccessDecision(False, "Acesso fora da faixa de horario permitida.")

    if not evaluate_time_rule(user_row.get("horario"), now.time()):
        return AccessDecision(False, "Acesso fora da faixa de horario permitida.")

    return AccessDecision(True)


def load_clients() -> list[dict[str, object]]:
    if not CLIENTS_FILE.exists():
        return []
    return workbook_rows(CLIENTS_FILE)


def search_clients(query: str, limit: int = 8) -> list[dict[str, object]]:
    query_text = normalize_text(query)
    query_digits = digits_only(query)
    if not query_text and not query_digits:
        return []

    matches: list[dict[str, object]] = []
    for client in load_clients():
        name = normalize_text(client.get("nome"))
        code = digits_only(client.get("codigo"))
        cnpj = digits_only(client.get("cnpj"))

        if query_digits and (code.startswith(query_digits) or cnpj.startswith(query_digits)):
            matches.append(client)
            continue
        if query_text and query_text in name:
            matches.append(client)

        if len(matches) >= limit:
            break
    return matches


def get_client_by_code(code: str) -> dict[str, object] | None:
    code_digits = digits_only(code)
    if not code_digits:
        return None
    for client in load_clients():
        if digits_only(client.get("codigo")) == code_digits:
            return client
    return None


def load_classification_reference() -> dict[str, list[str]]:
    if not CLASSIFICATION_FILE.exists():
        return {"canais": [], "tipologias": [], "segmentos": []}

    try:
        canal_rows = workbook_rows(CLASSIFICATION_FILE, "canal")
    except Exception:
        canal_rows = []

    try:
        tipologia_rows = workbook_rows(CLASSIFICATION_FILE, "tipologia")
    except Exception:
        tipologia_rows = []

    try:
        segmento_rows = workbook_rows(CLASSIFICATION_FILE, "segmento")
    except Exception:
        segmento_rows = []

    segmentos = [str(row.get("segmento") or "").strip() for row in segmento_rows if str(row.get("segmento") or "").strip()]
    if not segmentos:
        segmentos = [str(row.get("tipologia") or "").strip() for row in tipologia_rows if str(row.get("tipologia") or "").strip()]

    # Remove duplicados preservando ordem.
    seen: set[str] = set()
    segmentos_unicos: list[str] = []
    for segmento in segmentos:
        key = normalize_text(segmento)
        if key and key not in seen:
            seen.add(key)
            segmentos_unicos.append(segmento)

    return {
        "canais": [str(row.get("canal") or "").strip() for row in canal_rows if str(row.get("canal") or "").strip()],
        "tipologias": [str(row.get("tipologia") or "").strip() for row in tipologia_rows if str(row.get("tipologia") or "").strip()],
        "segmentos": segmentos_unicos,
    }


def append_submission(record: dict[str, object]) -> None:
    RESPONSES_FILE.parent.mkdir(parents=True, exist_ok=True)
    file_exists = RESPONSES_FILE.exists()
    with RESPONSES_FILE.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(record.keys()))
        if not file_exists:
            writer.writeheader()
        writer.writerow(record)


def _read_submissions_file_order() -> tuple[list[str], list[dict[str, str]]]:
    if not RESPONSES_FILE.exists():
        return [], []

    with RESPONSES_FILE.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    return fieldnames, rows


def load_submissions() -> list[dict[str, str]]:
    fieldnames, file_rows = _read_submissions_file_order()
    if not fieldnames:
        return []

    rows: list[dict[str, str]] = []
    for row_index, row in enumerate(file_rows, start=1):
        normalized_row = {key: str(value or "") for key, value in row.items()}
        normalized_row["_row_id"] = str(row_index)
        adjusted = normalized_row.get("qtd_checkouts_ajustado", "").strip()
        normalized_row["qtd_checkouts_final"] = adjusted if adjusted else normalized_row.get("qtd_checkouts", "")
        rows.append(normalized_row)

    rows.reverse()
    return rows


def get_submission_by_row_id(row_id: str) -> dict[str, str] | None:
    if not str(row_id).isdigit():
        return None

    target_index = int(row_id)
    if target_index <= 0:
        return None

    _, rows = _read_submissions_file_order()
    if target_index > len(rows):
        return None

    row = {key: str(value or "") for key, value in rows[target_index - 1].items()}
    row["_row_id"] = str(target_index)
    adjusted = row.get("qtd_checkouts_ajustado", "").strip()
    row["qtd_checkouts_final"] = adjusted if adjusted else row.get("qtd_checkouts", "")
    return row


def update_submission(row_id: str, updates: dict[str, object]) -> bool:
    if not str(row_id).isdigit():
        return False

    target_index = int(row_id)
    if target_index <= 0:
        return False

    fieldnames, rows = _read_submissions_file_order()
    if not fieldnames or target_index > len(rows):
        return False

    sanitized_updates = {str(key): str(value) for key, value in updates.items()}
    for key in sanitized_updates:
        if key not in fieldnames:
            fieldnames.append(key)

    target = rows[target_index - 1]
    for key, value in sanitized_updates.items():
        target[key] = value

    with RESPONSES_FILE.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})

    return True


def bucket_by_checkout_count(count: int) -> str:
    if count <= 4:
        return "1 a 4"
    if count <= 9:
        return "5 a 9"
    return "10+"
